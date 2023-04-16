# Last updated on: 2019-09-27

#!/usr/bin/env python
# test BLE Scanning software
# jcs 6/8/2014
import blescan
import sys
import bluetooth._bluetooth as bluez
import pymysql
from time import sleep
import threading
import os
import openpyxl
import time

import RPi.GPIO as GPIO
import time
import random

#LED_B = 24
#LED_G = 23
LED_R = 18

# GPIO 핀 모드 설정
GPIO.setmode(GPIO.BCM)
#GPIO.setup(LED_B, GPIO.OUT)
#GPIO.setup(LED_G, GPIO.OUT)
GPIO.setup(LED_R, GPIO.OUT)

# LED 모듈 색상 설정
#BLUE = GPIO.PWM(LED_B, 1000)
#GREEN = GPIO.PWM(LED_G, 1000)
RED = GPIO.PWM(LED_R, 1000)

# LED 모듈 초기 상태 설정
#BLUE.start(0)
#GREEN.start(0)
RED.start(0)

class DB_sending:
    def __init__(self):
        self.url = "210.115.227.108"
        self.id = 'cic'
        self.password = '20180903in'
        self.dbName = 'kindergartenbus'
    def creat_connet(self):
        self.db = pymysql.connect(host=self.url, port=3306, user=self.id, passwd=self.password, db=self.dbName, chars>
        self.cursor = self.db.cursor()

    def calcualte_distance_rssi(self, txPower, rssi):
        txPower_num =  int(txPower)
        rssi_num = int(rssi)
        if rssi_num ==  0 :
            return -1

        ratio = rssi_num * 1.0 / txPower_num
        if ratio < 1.0 :
            return str(ratio**10)
        else:
            distance = (0.89976) * (ratio**7.7095) + 0.111
            return str(distance)

    def insert_unique_data(self, mac, uuid, major, minor):
        sql = "insert into device_unique_info_tb (macAddress, UUID, major, minor) " \
                "select '"+ mac+"' ,'"+uuid+"' ,'"+major+"' ,'"+minor+"' from dual where not exists" \
                "( select * from device_unique_info_tb where macAddress = '"+mac+"' and UUID = '"+uuid+"')"
        print(sql)
        self.cursor.execute(sql)
        self.db.commit()
        print(self.cursor.lastrowid)

    def insert_valiable_data(self, mac, rssi, txpower, accuracy):
        sql = "INSERT INTO `device_variable_info_tb` (`macaddress`, `rssi`, `txpower`, `accuracy`, `time`) VALUES ('">
        print(sql)
        self.cursor.execute(sql)
        self.db.commit()
        print(self.cursor.lastrowid)

    def run_sensor_thread(self):
        os.system("sudo python3 /home/pi/sensorDataToDB.py")


dev_id = 0
conn = DB_sending()

try:
    sock = bluez.hci_open_dev(dev_id)
    print("ble thread started")

except:
    print("error accessing bluetooth device...")
    sys.exit(1)

blescan.hci_le_set_scan_parameters(sock)
blescan.hci_enable_le_scan(sock)


start_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))

wb = openpyxl.Workbook()
sheet1 = wb['Sheet']
sheet1.title = 'Collected data'
sheet1.cell(row=1, column=1).value = 'MAC'
sheet1.cell(row=1, column=2).value = 'Major'
sheet1.cell(row=1, column=3).value = 'Minor'
sheet1.cell(row=1, column=4).value = 'RSSI'
sheet1.cell(row=1, column=5).value = 'TX_power'
sheet1.cell(row=1, column=6).value = 'Accuracy'
sheet1.cell(row=1, column=7).value = 'yyyy-mm-dd h:m:s)'
wb.save(start_time + '.xlsx')

time_check = time.time()

while True:
        #print 'working?'
        returnedList = blescan.parse_events(sock, 10)
        #print 'working?2'
        for beacon in returnedList:
            #print(beacon)
            beacon_split = beacon.split(',')
            if beacon_split[3] in ["31206"]:
                sheet1.append([beacon_split[0], beacon_split[2], beacon_split[3], beacon_split[5], beacon_split[4], c>
                wb.save(start_time + '.xlsx')
                predicted_distance=beacon_split[5]
                predicted_distance = int(predicted_distance)
                print(predicted_distance)

        #print
                if predicted_distance <= -30 and predicted_distance > -50:
                    #for i in range(30):
                        RED.ChangeDutyCycle(100)  # LED 밝기 100%
                        time.sleep(0.1)  # 0.1초 대기
                        RED.ChangeDutyCycle(0)  # LED 끄기
                        time.sleep(0.1)  # 0.1초 대기
                elif predicted_distance <= -50 and predicted_distance > -65:
                    #for i in range(12):
                        RED.ChangeDutyCycle(50)  # LED 밝기 100%
                        time.sleep(0.4)  # 0.2초 대기
                        RED.ChangeDutyCycle(0)  # LED 끄기
                        time.sleep(0.4)  # 0.2초 대기
                elif predicted_distance <= -65 and predicted_distance >= -85:
                    #for i in range(4):
                        RED.ChangeDutyCycle(10)  # LED 밝기 100%
                        time.sleep(0.8)  # 0.5초 대기
                        RED.ChangeDutyCycle(0)  # LED 끄기
                        time.sleep(0.8)  # 0.5초 대기
                else:
                    #BLUE.ChangeDutyCycle(0)
                    #GREEN.ChangeDutyCycle(0)
                    RED.ChangeDutyCycle(0)

                #time.sleep(3) # 10초 주기로 예측 거리를 업데이트함
        print





#except KeyboardInterrupt:
#    #BLUE.stop()
#    #GREEN.stop()
#    RED.stop()
#    GPIO.cleanup()
